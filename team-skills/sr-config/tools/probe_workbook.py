# -*- coding: utf-8 -*-
"""probe_workbook.py — 策划配置工作簿标准探查工具（sr-config skill 附带）。

用法:
    python probe_workbook.py <工作簿路径> [--all-comments] [--max-rows N] [--json]

固定输出（不依赖 agent 当次是否"想到"，批注列永远在场）:
  1. 工作表清单与分类（运行时表 / INDEX / 其他）
  2. 每张运行时表的五行表头（中文字段/英文字段/简写/作用对象/数据结构）
  3. 每张运行时表的前 N 行样本数据（默认 3）
  4. 全部单元格批注（默认每表前 20 条；--all-comments 不限）——批注是项目字段契约
     的权威载体（枚举含义、codec 解析规则、取值边界），漏读批注 = 漏读契约
  5. INDEX / index 输出映射表全部内容
  6. 结构特性: 公式 / 合并单元格 / 数据验证 / Excel Table / 命名区域 / 隐藏行列 / 宏(.xlsm)

返回码: 0=正常; 2=文件不存在或不可读。
"""
import argparse
import json
import sys
from pathlib import Path

from openpyxl import load_workbook

HEADER_LABELS = ["中文字段", "英文字段", "简写", "作用对象", "数据结构"]


def is_runtime_sheet(ws) -> bool:
    """五行表头判据: A1=中文字段 且 A2=英文字段。分类仅供人工复核。"""
    a1 = ws.cell(1, 1).value
    if a1 is None or str(a1).strip() != HEADER_LABELS[0]:
        return False
    a2 = ws.cell(2, 1).value
    return a2 is not None and str(a2).strip() == HEADER_LABELS[1]


def comments_of(ws, limit):
    out = []
    for row in ws.iter_rows():
        for c in row:
            if c.comment is not None and c.comment.text:
                out.append({"cell": f"{ws.title}!{c.coordinate}", "text": c.comment.text.strip()})
                if limit and len(out) >= limit:
                    return out
    return out


def sheet_features(ws):
    feats = {
        "merged_cells": len(ws.merged_cells.ranges) if ws.merged_cells else 0,
        "hidden": bool(getattr(ws, "sheet_state", "visible") != "visible"),
        "data_validations": len(ws.data_validations.dataValidation) if getattr(ws, "data_validations", None) else 0,
        "tables": list(getattr(ws, "tables", {}) or {}),
        "formula_cells": 0,
    }
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.startswith("="):
                feats["formula_cells"] += 1
    return feats


def probe(path: Path, all_comments: bool, max_rows: int, as_json: bool):
    wb = load_workbook(path, data_only=False, keep_vba=path.suffix.lower() == ".xlsm")
    limit = None if all_comments else 20
    report = {"file": str(path), "sheets": [], "index": [], "warnings": []}

    index_sheet = next((s for s in wb.sheetnames if s.lower() == "index"), None)
    if index_sheet:
        for row in wb[index_sheet].iter_rows(values_only=True):
            if any(v is not None and str(v).strip() for v in row):
                report["index"].append([str(v) if v is not None else "" for v in row])

    for sn in wb.sheetnames:
        ws = wb[sn]
        entry = {
            "sheet": sn,
            "kind": "index" if sn == index_sheet else ("runtime" if is_runtime_sheet(ws) else "other"),
            "max_row": ws.max_row,
            "max_col": ws.max_column,
            "features": sheet_features(ws),
            "comments": comments_of(ws, limit),
            "warnings": [],
        }
        if entry["kind"] == "runtime":
            header = []
            for col in range(2, ws.max_column + 1):
                five = [ws.cell(r, col).value for r in range(1, 6)]
                if all(v is None for v in five):
                    break
                header.append({"col": ws.cell(1, col).column_letter, "zh": five[0],
                               "en": five[1], "short": five[2], "scope": five[3], "type": five[4]})
            entry["header"] = header
            samples = []
            for r in range(6, 6 + max_rows):
                vals = [ws.cell(r, c).value for c in range(2, ws.max_column + 1)]
                if all(v is None for v in vals):
                    break
                samples.append(vals)
            entry["samples"] = samples
            if not entry["comments"]:
                entry["warnings"].append("runtime_sheet_without_comments")
        report["sheets"].append(entry)

    if getattr(wb, "vba_archive", None):
        report["warnings"].append("contains_vba")

    if as_json:
        print(json.dumps(report, ensure_ascii=False, indent=1, default=str))
    else:
        print(f"=== {path.name} ===")
        for e in report["sheets"]:
            print(f"\n[{e['kind']}] {e['sheet']}  ({e['max_row']}x{e['max_col']})  feats={e['features']}")
            for w in e.get("warnings", []):
                print(f"  !! {w}")
            if e["kind"] == "runtime":
                print("  header:")
                for h in e["header"]:
                    print(f"    {h['col']:>3} {h['zh']} | {h['en']} | {h['short']} | {h['scope']} | {h['type']}")
                print("  samples:")
                for s in e["samples"]:
                    print("   ", s)
            if e["comments"]:
                shown = len(e["comments"])
                cap = "" if limit is None else f" (限{limit}条, --all-comments 看全部)"
                print(f"  comments ({shown}{cap}):")
                for c in e["comments"]:
                    txt = c["text"].replace("\n", " ⏎ ")
                    print(f"    {c['cell']}: {txt}")
            elif e["kind"] == "runtime":
                print("  comments: (无 — 若该表有字段契约规则, 此为证据缺口)")
        if report["index"]:
            print("\nINDEX:")
            for r in report["index"]:
                print("   ", r)
        if report["warnings"]:
            print("\nWARNINGS:", report["warnings"])
    return 0


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("workbook")
    ap.add_argument("--all-comments", action="store_true")
    ap.add_argument("--max-rows", type=int, default=3)
    ap.add_argument("--json", action="store_true")
    args = ap.parse_args()
    p = Path(args.workbook)
    if not p.is_file():
        print(f"文件不存在: {p}", file=sys.stderr)
        return 2
    return probe(p, args.all_comments, args.max_rows, args.json)


if __name__ == "__main__":
    sys.exit(main())
